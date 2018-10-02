from openpyxl import Workbook
from openpyxl import load_workbook
import logging 

def get_worksheet(workbook_path, sheet_index):
   wb = load_workbook(workbook_path, read_only=True)
   sheetname = wb.sheetnames[sheet_index]
   sheet = wb[sheetname]   
   return sheet

def read_worksheet(worksheet, headers, row_count, logger):
   row_index = 2
   total_values = []
   column_count = len(headers)   
   logger.debug("Total number of rows: " + str(row_count))
   logger.debug("Total number of columns: " + str(column_count))
   for row_index in range(2, row_count + 1):   # starting row_index from 2 because header line should be skipped
      row_cell_values = [] 
      for column_index in range(0, column_count):                
         logger.debug("Requesting1 row:" + str(row_index) + ", column:" + str(column_index) + ", max_row:" + str(worksheet.max_row) + ", max_column:" + str(worksheet.max_column))
         cell_value_text = get_single_cell_value(worksheet, row_index, column_index +1, logger)
         logger.debug("Cell text is: " + str(cell_value_text))

         if(not not cell_value_text):                 
            cell_value = {"header": headers[column_index]["header"], "column_index": headers[column_index]["index"], "value": cell_value_text }
            row_cell_values.append(cell_value)
            logger.debug("Added" + str(cell_value_text))
         

      if(len(row_cell_values) > 0):
         total_values.append({"RowNumber": row_index, "values": row_cell_values })
   return total_values
   
def get_headers(workbook_path, sheet_index, logger):
   headers = [] 
   header_column_index = 1
   sheet = get_worksheet(workbook_path,sheet_index)   
   headers = []
   logger.debug("")
   logger.debug("")
   logger.debug("Processing file: " + workbook_path)
   header = {"index": header_column_index, "header": get_single_cell_value(sheet, 1, header_column_index,logger)}
   logger.debug("Got first header")
   logger.debug("First header: "+ header["header"])
   while(not not header["header"]):
       logger.debug("Processing header: " + header["header"])   
       header_type = "string"	
       logger.debug("Going to read number of rows: " )		
       logger.debug("Number of rows: " + str(sheet.max_row))
       logger.debug("Going to read row 2, column " + str(header_column_index))
       first_row_cell_value =   get_single_cell_value(sheet, 2, header_column_index,logger)
       if(first_row_cell_value):
          logger.debug("Exception: could not get cell value, went for internal_value")		   		
          if(not not first_row_cell_value):
              header_type = type(first_row_cell_value)	
          header["type"]= {"Type": str(header_type)}
          logger.debug("Column type is: " + str(header_type))
          logger.debug("Processing file: " + workbook_path)	   
       headers.append(header)
       header_column_index+=1
       header = {"index": header_column_index, "header": sheet.cell(None, 1, header_column_index).value}
   return headers
   
def read_worksheet_content(filepath, headers, sheet_index, logger):
   worksheet = get_worksheet(filepath,0)   
   row_count = worksheet.max_row
   logger.debug("Max number of rows: " + str(row_count))
   config_file_content = read_worksheet(worksheet, headers, row_count, logger)
   return config_file_content   

def get_single_cell_value(worksheet, row_index, column_index, logger):   
   cell_value_text = "[Not assigned]"
   logger.debug(cell_value_text)
   logger.debug("Requesting2 row:" + str(row_index) + ", column:" + str(column_index) + ", max_row:" + str(worksheet.max_row) + ", max_column:" + str(worksheet.max_column))
   if(row_index<worksheet.max_row and column_index < worksheet.max_column+1):
      cell_value_text = worksheet.cell(row=row_index, column=column_index).value   
      logger.debug("Value read: "+str(cell_value_text))	   

   else:
       logger.debug("Cell out of bounds" + str(row_index))   
   if(not cell_value_text):
       cell_value_text = ""   
   logger.debug("Cell value: " + str(cell_value_text))
   return cell_value_text   